import os
import argparse
import openpyxl
from openpyxl.chart import SurfaceChart3D, Reference


parser = argparse.ArgumentParser()
parser.add_argument('open_file', type=str)
parser.add_argument('exel_file', type=str)
args = parser.parse_args()
open_file, exel_file = args.open_file, args.exel_file
is_exist = True
first = True
broken = False


def graphic(freq):
    chartsheet = workbook.create_chartsheet()
    chart = SurfaceChart3D()
    chart.height = 20
    chart.width = 30
    chart.title = freq
    labels = Reference(worksheet=datasheet, min_col=1, min_row=count_last, max_row=count)
    data = Reference(worksheet=datasheet, min_col=2, min_row=count_last, max_col=count - count_last, max_row=count)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chartsheet.add_chart(chart)


if os.path.exists(exel_file):
    workbook = openpyxl.load_workbook(exel_file)
else:
    is_exist = False
    workbook = openpyxl.Workbook()

worksheet = workbook["Sheet"]
datasheet = workbook.create_sheet("ChartData")
worksheet.cell(row=1, column=1).value = 'date'
worksheet.cell(row=1, column=2).value = 'time'
worksheet.cell(row=1, column=3).value = 'BatV'
worksheet.cell(row=1, column=4).value = 'BatT'
worksheet.cell(row=1, column=5).value = 'obsT'
worksheet.cell(row=1, column=6).value = 'lat'
worksheet.cell(row=1, column=7).value = 'lon'
worksheet.cell(row=1, column=8).value = 'alt'
worksheet.cell(row=1, column=9).value = 'RSSI'
count = 2

count_last = count
mp = 0
data = open(open_file, 'r')
for element in list(data):
    if ';' not in element:
        continue
    element = element.split(";")
    print(element[2])
    rx = element[-1]
    if first:
        last_rx = rx
    if count == 255 or rx != last_rx:
        graphic(last_rx)
        mp += 1
        count = 1
    last_rx = rx
    x, y, z = float(element[5]), float(element[6]), int(element[8])
    worksheet.cell(row=count + mp * 255, column=1).value = str(element[0])
    worksheet.cell(row=count + mp * 255, column=2).value = str(element[1])
    worksheet.cell(row=count + mp * 255, column=3).value = float(element[2])
    worksheet.cell(row=count + mp * 255, column=4).value = int(element[3])
    worksheet.cell(row=count + mp * 255, column=5).value = int(element[4])
    worksheet.cell(row=count + mp * 255, column=6).value = x
    for i in range(2, count + 1):
        if datasheet.cell(row=i, column=1).value == y and datasheet.cell(row=1, column=count).value == x and\
                datasheet.cell(row=i, count=i).value == z:
            broken = True
    if broken:
        continue
    datasheet.cell(row=count + 255 * mp, column=1).value = x
    datasheet.cell(row=255 * mp + 1, column=count).value = y
    datasheet.cell(row=count + 255 * mp, column=count).value = z
    worksheet.cell(row=count + mp * 255, column=7).value = y
    worksheet.cell(row=count + mp * 255, column=8).value = float(element[7])
    worksheet.cell(row=count + mp * 255, column=9).value = z
    count += 1
graphic(last_rx)


workbook.save("data.xlsx")
